# table_config_parser_service.py
import json
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from domains.table_config.table_config_model import TableConfig, ColumnConfig
from common.error import AppError
from domains.sql_generator.sql_generator_validator import SqlGeneratorValidator

TABLE_NAME_LABELS = {'наименование таблицы', 'table_name'}
COLUMN_CODE_LABELS = {'код колонки в бд', 'column_code'}
COLUMN_NAME_LABELS = {'описание', 'column_name'}
TYPE_LABELS = {'тип', 'type'}
SIZE_LABELS = {'размерность', 'size'}
REQUIRED_LABELS = {'обязательность', 'nullable'}
UNIQUE_LABELS = {'уникальность', 'unique'}
PRIMARY_KEY_LABELS = {'первичный ключ', 'primary_key'}
FOREIGN_KEY_LABELS = {'внешний ключ', 'foreign_key'}
DEFAULT_LABELS = {'значение по умолчанию', 'default'}


class TableConfigParserService:

    def __init__(self, validator: SqlGeneratorValidator | None = None) -> None:
        self._validator = validator or SqlGeneratorValidator()

    def parse_tables_config(self, content: bytes, filename: str) -> list[TableConfig]:
        extension = Path(filename).suffix.lower()
        if extension in {'.xlsx', '.xlsm'}:
            tables = self._parse_excel(content)
        elif extension == '.json':
            tables = self._parse_structured_tables(json.loads(content.decode('utf-8')))
        else:
            raise AppError('Неподдерживаемый формат файла.')

        if not tables:
            raise AppError('Секция tables_config не найдена или не содержит таблиц.')
        return tables

    def _parse_structured_tables(self, data: dict | None) -> list[TableConfig]:
        if not isinstance(data, dict):
            raise AppError('Некорректная структура конфигурации.')

        tables_config = data.get('tables_config')
        if tables_config is None:
            raise AppError('Секция tables_config отсутствует.')

        if isinstance(tables_config, dict):
            iterable = [{'table_name': k, 'columns': v} for k, v in tables_config.items()]
        elif isinstance(tables_config, list):
            iterable = tables_config
        else:
            raise AppError('Секция tables_config должна быть словарем или списком.')

        tables: list[TableConfig] = []
        for table_item in iterable:
            if not isinstance(table_item, dict):
                raise AppError('Каждая таблица в tables_config должна быть объектом.')
            table_name = self._normalize_text(table_item.get('table_name'))
            if not table_name:
                raise AppError('У таблицы отсутствует имя.')
            columns_raw = table_item.get('columns', [])
            if isinstance(columns_raw, dict):
                columns_raw = [{'column_code': name, **details} for name, details in columns_raw.items()]
            if not isinstance(columns_raw, list):
                raise AppError(f'Колонки таблицы {table_name} должны быть списком или словарем.')

            columns: list[ColumnConfig] = []
            for column_item in columns_raw:
                if not isinstance(column_item, dict):
                    raise AppError(f'Некорректное описание колонки в таблице {table_name}.')
                column_name = self._normalize_text(column_item.get('column_code'))
                db_type = self._normalize_text(column_item.get('type'))
                if not column_name or not db_type:
                    raise AppError(f'Колонка в таблице {table_name} должна содержать column_code и type.')
                columns.append(ColumnConfig(
                    name=column_name,
                    db_type=db_type,
                    size=self._normalize_text(column_item.get('size')),
                    nullable=self._to_bool(column_item.get('nullable', True), default=True),
                    unique=self._to_bool(column_item.get('unique', False), default=False),
                    primary_key=self._to_bool(column_item.get('primary_key', False), default=False),
                    foreign_key=self._normalize_text(column_item.get('foreign_key')),
                    default=self._normalize_text(column_item.get('default')),
                    label=self._normalize_text(column_item.get('column_name')),
                ))
            tables.append(TableConfig(name=table_name, columns=columns))

        return tables

    def _parse_excel(self, content: bytes) -> list[TableConfig]:
        workbook = load_workbook(BytesIO(content), data_only=True)

        v1_sheet = v2_sheet = None
        for sheet in workbook.worksheets:
            name = sheet.title.strip().lower()
            if name == 'tables_config':
                v1_sheet = sheet
            elif name == 'tables_config_v2':
                v2_sheet = sheet

        if v1_sheet is None and v2_sheet is None:
            if len(workbook.worksheets) == 1:
                sheet = workbook.worksheets[0]
                if self._is_v2_format(sheet):
                    v2_sheet = sheet
                else:
                    v1_sheet = sheet
            else:
                raise AppError('Лист tables_config не найден.')

        tables: list[TableConfig] = []
        if v1_sheet is not None:
            tables.extend(self._parse_excel_v1_rows([list(r) for r in v1_sheet.iter_rows(values_only=True)]))
        if v2_sheet is not None:
            tables.extend(self._parse_excel_v2_sheet(v2_sheet))
        return tables

    def _is_v2_format(self, sheet) -> bool:
        rows = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        if not rows:
            return False
        for cell in rows[0]:
            label = self._label(cell)
            if label in COLUMN_NAME_LABELS or label in COLUMN_CODE_LABELS:
                return True
        return False

    def _parse_excel_v1_rows(self, rows: list[list]) -> list[TableConfig]:
        tables: list[TableConfig] = []
        row_index = 0
        while row_index < len(rows):
            current_label = self._label(rows[row_index][0] if rows[row_index] else None)
            if current_label in TABLE_NAME_LABELS:
                table, row_index = self._parse_excel_table_block(rows, row_index)
                if table is not None:
                    tables.append(table)
                continue
            row_index += 1
        return tables

    def _parse_excel_v2_sheet(self, sheet) -> list[TableConfig]:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        if len(rows) < 2:
            return []

        table_row, header_row = rows[0], rows[1]
        block_starts = [i for i, cell in enumerate(header_row) if self._label(cell) in COLUMN_NAME_LABELS]

        tables: list[TableConfig] = []
        for i, start_col in enumerate(block_starts):
            raw_name_cell = table_row[start_col] if start_col < len(table_row) else None
            if self._label(raw_name_cell) in TABLE_NAME_LABELS:
                raw_name_cell = table_row[start_col + 1] if start_col + 1 < len(table_row) else None
            table_name = self._normalize_text(raw_name_cell)
            if not table_name:
                continue

            end_col = block_starts[i + 1] if i + 1 < len(block_starts) else len(header_row)
            header_map: dict[str, int] = {}
            for col_idx in range(start_col, min(end_col, len(header_row))):
                label = self._label(header_row[col_idx])
                if label and label not in header_map:
                    header_map[label] = col_idx

            code_col = self._find_col(header_map, COLUMN_CODE_LABELS)
            type_col = self._find_col(header_map, TYPE_LABELS)
            if code_col is None or type_col is None:
                raise AppError(f'Для таблицы {table_name} отсутствуют заголовки кодов колонок или типов.')

            name_col = self._find_col(header_map, COLUMN_NAME_LABELS)
            size_col = self._find_col(header_map, SIZE_LABELS)
            required_col = self._find_col(header_map, REQUIRED_LABELS)
            unique_col = self._find_col(header_map, UNIQUE_LABELS)
            pk_col = self._find_col(header_map, PRIMARY_KEY_LABELS)
            fk_col = self._find_col(header_map, FOREIGN_KEY_LABELS)
            default_col = self._find_col(header_map, DEFAULT_LABELS)

            columns: list[ColumnConfig] = []
            for row in rows[2:]:
                code = self._cell(row, code_col)
                if not code:
                    continue
                db_type = self._cell(row, type_col)
                if not db_type:
                    raise AppError(f'У колонки {code} таблицы {table_name} не указан тип.')

                fk_value = self._cell(row, fk_col) if fk_col is not None else None
                req_val = self._cell(row, required_col) if required_col is not None else None
                uniq_val = self._cell(row, unique_col) if unique_col is not None else None
                pk_val = self._cell(row, pk_col) if pk_col is not None else None

                self._validator.validate_yes_no_cell(req_val, 'Обязательность', code, table_name)
                self._validator.validate_yes_no_cell(uniq_val, 'Уникальность', code, table_name)
                self._validator.validate_yes_no_cell(pk_val, 'Первичный ключ', code, table_name)
                self._validator.validate_reference_cell(fk_value, code, table_name)

                columns.append(ColumnConfig(
                    name=code, db_type=db_type,
                    size=self._cell(row, size_col) if size_col is not None else None,
                    nullable=not self._contains_any(req_val, {'да'}),
                    unique=self._contains_any(uniq_val, {'да'}),
                    primary_key=self._contains_any(pk_val, {'да'}),
                    foreign_key=fk_value,
                    default=self._cell(row, default_col) if default_col is not None else None,
                    label=self._cell(row, name_col) if name_col is not None else None,
                ))
            tables.append(TableConfig(name=table_name, columns=columns))

        return tables

    def _parse_excel_table_block(self, rows: list[list], start_index: int) -> tuple[TableConfig | None, int]:
        header_row = rows[start_index]
        table_name = self._first_non_empty(header_row[1:])
        if not table_name:
            return None, start_index + 1

        end_index = start_index + 1
        row_map: dict[str, list] = {}
        while end_index < len(rows):
            row = rows[end_index]
            label = self._label(row[0] if row else None)
            if label in TABLE_NAME_LABELS:
                break
            if label:
                row_map[label] = row
            end_index += 1

        code_row = self._find_row(row_map, COLUMN_CODE_LABELS)
        type_row = self._find_row(row_map, TYPE_LABELS)
        if not code_row or not type_row:
            raise AppError(f'Для таблицы {table_name} отсутствуют строки с кодами колонок или типами.')

        name_row = self._find_row(row_map, COLUMN_NAME_LABELS)
        size_row = self._find_row(row_map, SIZE_LABELS)
        required_row = self._find_row(row_map, REQUIRED_LABELS)
        unique_row = self._find_row(row_map, UNIQUE_LABELS)
        primary_key_row = self._find_row(row_map, PRIMARY_KEY_LABELS)
        foreign_key_row = self._find_row(row_map, FOREIGN_KEY_LABELS)
        default_row = self._find_row(row_map, DEFAULT_LABELS)

        max_len = max(len(r) for r in [code_row, type_row] + [x for x in [
            name_row, size_row, required_row, unique_row, primary_key_row, foreign_key_row, default_row
        ] if x])

        columns: list[ColumnConfig] = []
        for idx in range(1, max_len):
            name = self._cell(code_row, idx)
            db_type = self._cell(type_row, idx)
            if not name:
                continue
            if not db_type:
                raise AppError(f'У колонки {name} таблицы {table_name} не указан тип.')

            foreign_key_value = self._cell(foreign_key_row, idx)
            self._validator.validate_yes_no_cell(self._cell(required_row, idx), 'Обязательность', name, table_name)
            self._validator.validate_yes_no_cell(self._cell(unique_row, idx), 'Уникальность', name, table_name)
            self._validator.validate_yes_no_cell(self._cell(primary_key_row, idx), 'Первичный ключ', name, table_name)
            self._validator.validate_reference_cell(foreign_key_value, name, table_name)

            columns.append(ColumnConfig(
                name=name, db_type=db_type,
                size=self._cell(size_row, idx),
                nullable=not self._contains_any(self._cell(required_row, idx), {'да'}),
                unique=self._contains_any(self._cell(unique_row, idx), {'да'}),
                primary_key=self._contains_any(self._cell(primary_key_row, idx), {'да'}),
                foreign_key=foreign_key_value,
                default=self._cell(default_row, idx),
                label=self._cell(name_row, idx),
            ))

        return TableConfig(name=table_name, columns=columns), end_index

    @staticmethod
    def _find_row(row_map: dict[str, list], names: set[str]) -> list | None:
        for name in names:
            if name in row_map:
                return row_map[name]
        return None

    @staticmethod
    def _find_col(header_map: dict[str, int], labels: set[str]) -> int | None:
        for label in labels:
            if label in header_map:
                return header_map[label]
        return None

    def _cell(self, row: list | None, index: int) -> str | None:
        if not row or index >= len(row):
            return None
        return self._normalize_text(row[index])

    def _label(self, value) -> str:
        return self._normalize_text(value, lower=True) or ''

    @staticmethod
    def _normalize_text(value, lower: bool = False) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        return text.lower() if lower else text

    def _first_non_empty(self, values: list) -> str | None:
        for value in values:
            text = self._normalize_text(value)
            if text:
                return text
        return None

    @staticmethod
    def _contains_any(value: str | None, expected_values: set[str]) -> bool:
        if not value:
            return False
        return any(item in value.lower() for item in expected_values)

    @staticmethod
    def _to_bool(value, default: bool) -> bool:
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        normalized = str(value).strip().lower()
        if normalized in {'1', 'true', 'yes', 'y', 'да'}:
            return True
        if normalized in {'0', 'false', 'no', 'n', 'нет'}:
            return False
        return default
