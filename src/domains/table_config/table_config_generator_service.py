#table_config_generator_service.py
import re
import zipfile
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

from flask import Request, Response
from openpyxl import load_workbook

from common.error import AppError
from common.project_paths import ProjectPaths
from domains.table_config.table_config_data_file_reader_service import (
    ALLOWED_DATA_EXTENSIONS,
    TableConfigDataFileReaderService,
)
from utils.file_util import read_uploaded_file

_TEMPLATE_PATH = ProjectPaths.STATIC / 'TablesConfig.xlsm'
_V2_DATA_COLS = 9

_AUTO_PK_COL = {'label': 'Ид', 'code': 'id', 'db_type': 'bigserial', 'primary_key': True}
_PACKAGE_ID_COL = {'label': 'Пакетный ид', 'code': 'package_id', 'db_type': 'varchar'}
_PACKAGE_TS_COL = {'label': 'Пакетный временной штамп', 'code': 'package_timestamp', 'db_type': 'timestamptz'}


class TableConfigGeneratorService:

    def __init__(self, reader: TableConfigDataFileReaderService | None = None) -> None:
        self._reader = reader or TableConfigDataFileReaderService()

    def generate_table_config_from_data_file(self, request: Request) -> Response:
        content, filename = read_uploaded_file(request.files.get('data_file'), ALLOWED_DATA_EXTENSIONS)
        add_pk = request.form.get('add_pk') == '1'
        add_package_fields = request.form.get('add_package_fields') == '1'

        try:
            table_name, headers, rows = self._reader.read_data_file(content, filename)
            columns = self._reader.infer_columns(headers, rows)
            xlsx_bytes = self.generate_excel_config_v2(table_name, columns, add_pk=add_pk, add_package_fields=add_package_fields)
        except AppError:
            raise
        except Exception as exc:
            raise AppError(str(exc)) from exc

        download_name = f'{table_name}_config.xlsm'
        ascii_name = download_name.encode('ascii', 'replace').decode('ascii')
        encoded_name = quote(download_name, encoding='utf-8')
        return Response(
            xlsx_bytes,
            mimetype='application/vnd.ms-excel.sheet.macroEnabled.12',
            headers={
                'Content-Disposition': (
                    f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{encoded_name}"
                ),
            },
        )

    def generate_excel_config_v2(
        self,
        table_name: str,
        columns: list[dict],
        add_pk: bool = False,
        add_package_fields: bool = False,
    ) -> bytes:
        columns = list(columns)

        if add_pk and not any(c.get('primary_key') for c in columns):
            columns = [_AUTO_PK_COL] + columns

        if add_package_fields:
            existing = {c['code'].lower() for c in columns}
            need_pkg_id = 'package_id' not in existing
            need_pkg_ts = 'package_timestamp' not in existing
            if need_pkg_id or need_pkg_ts:
                ref_idx = next((i for i, c in enumerate(columns) if c['code'].lower() == 'package_id'), -1)
                if ref_idx < 0:
                    ref_idx = next((i for i, c in enumerate(columns) if c['code'].lower() == 'id'), -1)
                insert_at = ref_idx + 1 if ref_idx >= 0 else 0
                pkg_cols = []
                if need_pkg_id:
                    pkg_cols.append(_PACKAGE_ID_COL)
                if need_pkg_ts:
                    pkg_cols.append(_PACKAGE_TS_COL)
                columns = columns[:insert_at] + pkg_cols + columns[insert_at:]

        wb = load_workbook(_TEMPLATE_PATH, keep_vba=True)
        template_bytes = _TEMPLATE_PATH.read_bytes()
        ws = wb['tables_config_v2']
        wb.active = ws

        ws.cell(row=1, column=2).value = table_name
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=_V2_DATA_COLS):
            for cell in row:
                cell.value = None

        for row_idx, col_info in enumerate(columns, start=3):
            ws.cell(row=row_idx, column=1).value = col_info.get('label') or col_info['code']
            ws.cell(row=row_idx, column=2).value = col_info['code']
            ws.cell(row=row_idx, column=3).value = col_info['db_type']
            if col_info.get('size'):
                ws.cell(row=row_idx, column=4).value = col_info['size']
            if col_info.get('primary_key'):
                ws.cell(row=row_idx, column=7).value = 'да'

        for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=_V2_DATA_COLS):
            max_len = max((len(str(cell.value)) for cell in col_cells if cell.value is not None), default=0)
            if max_len > 0:
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

        output = BytesIO()
        wb.save(output)
        return self._restore_x14_validations(template_bytes, output.getvalue())

    @staticmethod
    def _sheet_name_to_zip_path(workbook_xml: str, rels_xml: str) -> dict[str, str]:
        name_rid: dict[str, str] = {}
        for m in re.finditer(r'<sheet\s[^>]*\bname="([^"]*)"[^>]*\br:id="([^"]*)"', workbook_xml):
            name_rid[m.group(1)] = m.group(2)
        rid_zip: dict[str, str] = {}
        for m in re.finditer(r'<Relationship\s([^>]*)>', rels_xml):
            attrs = m.group(1)
            id_m = re.search(r'\bId="([^"]*)"', attrs)
            target_m = re.search(r'\bTarget="([^"]*)"', attrs)
            if not id_m or not target_m:
                continue
            target = target_m.group(1)
            zip_path = target.lstrip('/') if target.startswith('/') else f'xl/{target}'
            rid_zip[id_m.group(1)] = zip_path
        return {name: rid_zip[rid] for name, rid in name_rid.items() if rid in rid_zip}

    def _restore_x14_validations(self, template_bytes: bytes, output_bytes: bytes) -> bytes:
        wb_rel_path = 'xl/_rels/workbook.xml.rels'
        wb_path = 'xl/workbook.xml'

        with zipfile.ZipFile(BytesIO(template_bytes)) as tz:
            tmpl_name_to_zip = self._sheet_name_to_zip_path(
                tz.read(wb_path).decode('utf-8'),
                tz.read(wb_rel_path).decode('utf-8'),
            )
            tmpl_extlst: dict[str, tuple[str, str] | None] = {}
            for name, zip_path in tmpl_name_to_zip.items():
                xml = tz.read(zip_path).decode('utf-8')
                m = re.search(r'<extLst>.*?</extLst>', xml, re.DOTALL)
                if not m:
                    tmpl_extlst[name] = None
                    continue
                root_m = re.search(r'<worksheet([^>]*)>', xml)
                tmpl_extlst[name] = (m.group(0), root_m.group(1) if root_m else '')

        out_buf = BytesIO(output_bytes)
        with zipfile.ZipFile(out_buf, 'r') as oz:
            out_name_to_zip = self._sheet_name_to_zip_path(
                oz.read(wb_path).decode('utf-8'),
                oz.read(wb_rel_path).decode('utf-8'),
            )
            out_entries = {item.filename: oz.read(item.filename) for item in oz.infolist()}

        for sheet_name, payload in tmpl_extlst.items():
            if not payload:
                continue
            extlst, tmpl_root_attrs = payload
            zip_path = out_name_to_zip.get(sheet_name)
            if not zip_path or zip_path not in out_entries:
                continue
            xml = out_entries[zip_path].decode('utf-8')
            root_tag_m = re.match(r'(<worksheet)([^>]*)(>)', xml)
            if root_tag_m:
                cur_attrs = root_tag_m.group(2)
                extra_ns = ''
                for ns_m in re.finditer(r'(xmlns:[a-zA-Z0-9_]+="[^"]*")', tmpl_root_attrs):
                    ns_decl = ns_m.group(1)
                    prefix = re.match(r'xmlns:([a-zA-Z0-9_]+)=', ns_decl).group(1)
                    if (f'{prefix}:' in extlst) and (f'xmlns:{prefix}=' not in cur_attrs):
                        extra_ns += ' ' + ns_decl
                if extra_ns:
                    xml = root_tag_m.group(1) + cur_attrs + extra_ns + root_tag_m.group(3) + xml[root_tag_m.end():]
            xml = re.sub(r'<extLst>.*?</extLst>', '', xml, flags=re.DOTALL)
            xml = xml.replace('</worksheet>', extlst + '</worksheet>', 1)
            out_entries[zip_path] = xml.encode('utf-8')

        result = BytesIO()
        with zipfile.ZipFile(result, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
            for filename, data in out_entries.items():
                zout.writestr(filename, data)
        return result.getvalue()
