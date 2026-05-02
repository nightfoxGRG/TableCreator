# table_config_data_file_reader_service.py
"""Чтение файла данных и формирование промежуточной структуры данных для последующего формирования файла конфигурации

Supports reading Excel (.xlsx, .xlsm) and CSV files.
The first row of the input file is expected to contain column headers.
"""

import csv
import re
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook

from domains.libretranslate.libretranslate_service import LibreTranslateService

ALLOWED_DATA_EXTENSIONS = {'.xlsx', '.xlsm', '.csv'}

_DATE_RE = re.compile(r'^\d{4}-\d{2}-\d{2}$')
_DATETIME_RE = re.compile(r'^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}')
_BOOL_VALUES: frozenset[str] = frozenset({'true', 'false'})


class TableConfigDataFileReaderService:

    def __init__(self, libretranslate: LibreTranslateService | None = None) -> None:
        self._libretranslate = libretranslate or LibreTranslateService()

    def read_data_file(self, content: bytes, filename: str) -> tuple[str, list[str], list[list]]:
        """Parse *content* from *filename* and return (table_name, headers, rows)."""
        extension = Path(filename).suffix.lower()
        stem = Path(filename).stem

        if extension in {'.xlsx', '.xlsm'}:
            headers, rows = self._read_excel(content)
        elif extension == '.csv':
            headers, rows = self._read_csv(content)
        else:
            from common.error import AppError
            raise AppError(
                f'Неподдерживаемый формат файла данных. '
                f'Допустимы: {", ".join(sorted(ALLOWED_DATA_EXTENSIONS))}.'
            )

        return stem, headers, rows

    def infer_columns(self, headers: list[str], rows: list[list]) -> list[dict]:
        """Return a list of column-info dicts inferred from *headers* and *rows*."""
        columns = []
        for col_idx, header in enumerate(headers):
            values = [row[col_idx] if col_idx < len(row) else None for row in rows]
            db_type, size = self._infer_db_type(values)
            columns.append({
                'code': self._sanitize_code(header),
                'label': header,
                'db_type': db_type,
                'size': size,
            })
        return columns

    @staticmethod
    def _read_excel(content: bytes) -> tuple[list[str], list[list]]:
        wb = load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
        all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
        if not all_rows:
            from common.error import AppError
            raise AppError('Файл данных пустой.')
        headers = [str(cell) if cell is not None else '' for cell in all_rows[0]]
        return headers, all_rows[1:]

    @staticmethod
    def _read_csv(content: bytes) -> tuple[list[str], list[list]]:
        for encoding in ('utf-8-sig', 'utf-8', 'cp1251', 'latin-1'):
            try:
                text = content.decode(encoding)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        else:
            text = content.decode('latin-1')

        reader = csv.reader(StringIO(text))
        all_rows = list(reader)
        if not all_rows:
            from common.error import AppError
            raise AppError('Файл данных пустой.')
        return all_rows[0], all_rows[1:]

    def _sanitize_code(self, name: str) -> str:
        code = self._libretranslate.translate_to_english(name).strip()
        code = re.sub(r"[\s\-']+", '_', code)
        code = re.sub(r'[^A-Za-z0-9_]', '', code)
        if code and code[0].isdigit():
            code = '_' + code
        return code.lower() or 'col'

    def _infer_db_type(self, values: list) -> tuple[str, str | None]:
        nn = self._non_null(values)

        if not nn:
            return 'text', None
        if all(self._is_bool(v) for v in nn):
            return 'boolean', None
        if all(self._is_date(v) or self._is_datetime(v) for v in nn):
            return ('timestamp', None) if any(self._is_datetime(v) for v in nn) else ('date', None)
        if all(self._try_int(v) for v in nn):
            max_val = max(abs(int(str(v).strip())) for v in nn)
            return ('bigint', None) if max_val > 2_147_483_647 else ('integer', None)
        if all(self._try_float(v) for v in nn):
            max_int_digits = max_dec_digits = 0
            for v in nn:
                s = str(float(str(v).strip()))
                if '.' in s:
                    int_part, dec_part = s.split('.')
                    max_int_digits = max(max_int_digits, len(int_part.lstrip('-')))
                    max_dec_digits = max(max_dec_digits, len(dec_part.rstrip('0')))
                else:
                    max_int_digits = max(max_int_digits, len(s.lstrip('-')))
            if max_dec_digits > 0:
                return 'numeric', f'{max_int_digits + max_dec_digits},{max_dec_digits}'
            return 'numeric', None

        max_len = max(len(str(v)) for v in nn)
        if max_len > 255:
            return 'text', None
        return 'varchar', str(self._round_up_to_50(max_len))

    @staticmethod
    def _non_null(values: list) -> list:
        return [v for v in values if v is not None and str(v).strip() != '']

    @staticmethod
    def _try_int(v) -> bool:
        try:
            int(str(v).strip())
            return True
        except (ValueError, TypeError):
            return False

    @staticmethod
    def _try_float(v) -> bool:
        try:
            float(str(v).strip())
            return True
        except (ValueError, TypeError):
            return False

    @staticmethod
    def _is_bool(v) -> bool:
        return str(v).strip().lower() in _BOOL_VALUES

    @staticmethod
    def _is_date(v) -> bool:
        return bool(_DATE_RE.match(str(v).strip()))

    @staticmethod
    def _is_datetime(v) -> bool:
        return bool(_DATETIME_RE.match(str(v).strip()))

    @staticmethod
    def _round_up_to_50(n: int) -> int:
        return max(50, ((n + 49) // 50) * 50)
