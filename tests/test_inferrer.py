import csv
import io
import re
from unittest.mock import patch, MagicMock

import pytest
from openpyxl import Workbook

from services.inferrer import (
    ALLOWED_DATA_EXTENSIONS,
    _round_up_to_50,
    _sanitize_code,
    infer_columns,
    read_data_file,
)
from services.config_generator import generate_excel_config_v2
from services.models import ConfigParseError
from app import create_app


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_excel_bytes(headers: list, rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_csv_bytes(headers: list, rows: list[list], encoding: str = 'utf-8') -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode(encoding)


# ---------------------------------------------------------------------------
# _round_up_to_50
# ---------------------------------------------------------------------------

def test_round_up_to_50_below_50():
    assert _round_up_to_50(1) == 50
    assert _round_up_to_50(50) == 50


def test_round_up_to_50_above_50():
    assert _round_up_to_50(51) == 100
    assert _round_up_to_50(100) == 100
    assert _round_up_to_50(101) == 150


# ---------------------------------------------------------------------------
# _sanitize_code
# ---------------------------------------------------------------------------

def test_sanitize_code_spaces():
    assert _sanitize_code('First Name') == 'first_name'


def test_sanitize_code_special_chars():
    assert _sanitize_code('col#1!') == 'col1'


def test_sanitize_code_leading_digit():
    assert _sanitize_code('1col') == '_1col'


def test_sanitize_code_empty():
    assert _sanitize_code('') == 'col'


def test_sanitize_code_cyrillic():
    mock_response = MagicMock()
    mock_response.json.return_value = {'translatedText': 'User name'}
    mock_response.raise_for_status.return_value = None
    with patch('services.inferrer.requests.post', return_value=mock_response):
        code = _sanitize_code('Имя пользователя')
    # Must be non-empty and contain only lowercase ASCII-identifier chars
    assert code and code != 'col'
    assert re.match(r'^[a-z][a-z0-9_]*$', code), f'Not a valid identifier: {code!r}'


def test_sanitize_code_cyrillic_single_word():
    mock_response = MagicMock()
    mock_response.json.return_value = {'translatedText': 'Identifier'}
    mock_response.raise_for_status.return_value = None
    with patch('services.inferrer.requests.post', return_value=mock_response):
        code = _sanitize_code('Идентификатор')
    assert code == 'identifier'


# ---------------------------------------------------------------------------
# read_data_file — Excel
# ---------------------------------------------------------------------------

def test_read_data_file_excel():
    content = _make_excel_bytes(
        ['id', 'name', 'age'],
        [[1, 'Alice', 30], [2, 'Bob', 25]],
    )
    table_name, headers, rows = read_data_file(content, 'users.xlsx')
    assert table_name == 'users'
    assert headers == ['id', 'name', 'age']
    assert len(rows) == 2


def test_read_data_file_csv():
    content = _make_csv_bytes(
        ['product', 'price'],
        [['Apple', '1.50'], ['Banana', '0.75']],
    )
    table_name, headers, rows = read_data_file(content, 'products.csv')
    assert table_name == 'products'
    assert headers == ['product', 'price']
    assert len(rows) == 2


def test_read_data_file_unsupported_extension_raises():
    with pytest.raises(ConfigParseError, match='Неподдерживаемый формат'):
        read_data_file(b'data', 'file.txt')


# ---------------------------------------------------------------------------
# infer_columns — type detection
# ---------------------------------------------------------------------------

def test_infer_integer_column():
    headers = ['qty']
    rows = [[1], [2], [100], [None]]
    cols = infer_columns(headers, rows)
    assert cols[0]['db_type'] == 'integer'
    assert cols[0]['size'] is None


def test_infer_bigint_column():
    headers = ['big_id']
    rows = [[3_000_000_000]]
    cols = infer_columns(headers, rows)
    assert cols[0]['db_type'] == 'bigint'


def test_infer_numeric_column():
    headers = ['price']
    rows = [['1.99'], ['0.50'], ['123.45']]
    cols = infer_columns(headers, rows)
    assert cols[0]['db_type'] == 'numeric'
    assert cols[0]['size'] is not None


def test_infer_boolean_column():
    headers = ['flag']
    rows = [['true'], ['false'], ['true']]
    cols = infer_columns(headers, rows)
    assert cols[0]['db_type'] == 'boolean'


def test_infer_date_column():
    headers = ['birth_date']
    rows = [['2000-01-01'], ['1990-12-31']]
    cols = infer_columns(headers, rows)
    assert cols[0]['db_type'] == 'date'


def test_infer_timestamp_column():
    headers = ['created_at']
    rows = [['2024-01-15 10:30:00'], ['2024-06-01T08:00:00']]
    cols = infer_columns(headers, rows)
    assert cols[0]['db_type'] == 'timestamp'


def test_infer_varchar_column():
    headers = ['code']
    rows = [['ABC'], ['XYZ'], ['HELLO']]
    cols = infer_columns(headers, rows)
    assert cols[0]['db_type'] == 'varchar'
    assert cols[0]['size'] == '50'


def test_infer_text_column():
    headers = ['description']
    rows = [['x' * 300]]
    cols = infer_columns(headers, rows)
    assert cols[0]['db_type'] == 'text'
    assert cols[0]['size'] is None


def test_infer_empty_column_defaults_to_text():
    headers = ['notes']
    rows = [[None], [''], [None]]
    cols = infer_columns(headers, rows)
    assert cols[0]['db_type'] == 'text'


def test_infer_columns_preserves_label_and_code():
    headers = ['First Name']
    rows = [['Alice']]
    cols = infer_columns(headers, rows)
    assert cols[0]['label'] == 'First Name'
    assert cols[0]['code'] == 'first_name'


# ---------------------------------------------------------------------------
# generate_excel_config_v2
# ---------------------------------------------------------------------------

def test_generate_excel_config_v2_roundtrip():
    """Generated file can be re-opened and contains correct sheet/headers."""
    from openpyxl import load_workbook

    columns = [
        {'code': 'id', 'label': 'Идентификатор', 'db_type': 'bigserial', 'size': None},
        {'code': 'name', 'label': 'Имя', 'db_type': 'varchar', 'size': '100'},
        {'code': 'score', 'label': 'Баллы', 'db_type': 'integer', 'size': None},
    ]
    xlsx_bytes = generate_excel_config_v2('players', columns)

    wb = load_workbook(io.BytesIO(xlsx_bytes), keep_vba=True)
    assert 'tables_config_v2' in wb.sheetnames

    ws = wb['tables_config_v2']
    # Row 1: label in A1, table name in B1
    assert ws.cell(row=1, column=1).value == 'Наименование таблицы'
    assert ws.cell(row=1, column=2).value == 'players'
    # Row 2 = headers
    assert ws.cell(row=2, column=2).value == 'Код колонки в БД'
    # Row 3 = first column
    assert ws.cell(row=3, column=2).value == 'id'
    assert ws.cell(row=3, column=3).value == 'bigserial'
    # Row 4 = second column has size
    assert ws.cell(row=4, column=4).value == '100'


def test_generate_excel_config_v2_full_pipeline():
    """Full pipeline: data file → infer → generate config → re-parse with parser."""
    from services.parser import parse_tables_config

    content = _make_csv_bytes(
        ['user_id', 'username', 'score'],
        [['1', 'alice', '42'], ['2', 'bob', '99']],
    )
    table_name, headers, rows = read_data_file(content, 'stats.csv')
    columns = infer_columns(headers, rows)
    xlsx_bytes = generate_excel_config_v2(table_name, columns)

    tables = parse_tables_config(xlsx_bytes, 'stats_config.xlsx')
    assert len(tables) == 1
    assert tables[0].name == 'stats'
    col_names = [c.name for c in tables[0].columns]
    assert 'user_id' in col_names
    assert 'username' in col_names
    assert 'score' in col_names


# ---------------------------------------------------------------------------
# Web route: Russian (Cyrillic) filename must not hang the server
# ---------------------------------------------------------------------------

def test_inferrer_generate_cyrillic_filename():
    """Uploading a file with a Cyrillic name must return a valid response.

    Previously the Content-Disposition header was built with a raw non-ASCII
    filename, which caused Werkzeug to fail encoding the header as Latin-1 and
    the response was never sent (browser appeared to hang).
    """
    app = create_app()
    client = app.test_client()

    csv_bytes = _make_csv_bytes(
        ['id', 'название'],
        [['1', 'Тест'], ['2', 'Данные']],
    )

    data = {
        'data_file': (io.BytesIO(csv_bytes), 'Данные.csv'),
    }
    response = client.post(
        '/inferrer/generate',
        data=data,
        content_type='multipart/form-data',
    )

    assert response.status_code == 200
    cd = response.headers.get('Content-Disposition', '')
    assert 'attachment' in cd
    # RFC 5987 encoded name must be present
    assert "filename*=UTF-8''" in cd
    # The Cyrillic characters must appear percent-encoded, not raw
    assert '%' in cd
