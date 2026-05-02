from io import BytesIO

import pytest
from openpyxl import Workbook

from common.error import AppError
from domains.table_config.table_config_parser_service import TableConfigParserService

_parser = TableConfigParserService()


def test_parse_excel_tables_config_like_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'tables_config'

    ws['A1'] = 'Наименование таблицы'
    ws['B1'] = 'test_table'
    ws['A2'] = 'Описание'
    ws['B2'] = 'Идентификатор'
    ws['C2'] = 'Код'
    ws['D2'] = 'ID другой таблицы'
    ws['A3'] = 'Код колонки в БД'
    ws['B3'] = 'id'
    ws['C3'] = 'code'
    ws['D3'] = 'another_table_id'
    ws['A4'] = 'Тип'
    ws['B4'] = 'bigserial'
    ws['C4'] = 'varchar'
    ws['D4'] = 'bigint'
    ws['A5'] = 'Размерность'
    ws['C5'] = '50'
    ws['A6'] = 'Обязательность'
    ws['B6'] = 'да'
    ws['C6'] = 'да'
    ws['D6'] = 'да'
    ws['A8'] = 'Первичный ключ'
    ws['B8'] = 'да'
    ws['A9'] = 'Внешний ключ'
    ws['D9'] = 'another_table(id)'

    payload = BytesIO()
    wb.save(payload)

    tables = _parser.parse_tables_config(payload.getvalue(), 'config.xlsx')

    assert len(tables) == 1
    assert tables[0].name == 'test_table'
    assert [c.name for c in tables[0].columns] == ['id', 'code', 'another_table_id']
    assert tables[0].columns[1].size == '50'
    assert tables[0].columns[0].primary_key is True
    assert tables[0].columns[2].foreign_key == 'another_table(id)'
    assert tables[0].columns[0].label == 'Идентификатор'
    assert tables[0].columns[1].label == 'Код'
    assert tables[0].columns[2].label == 'ID другой таблицы'


# ---------------------------------------------------------------------------
# v2 format tests (tables_config_v2 sheet — horizontal / tabular layout)
# ---------------------------------------------------------------------------

def _make_v2_workbook(tables: list[dict]) -> bytes:
    """Build a workbook with a tables_config_v2 sheet.

    Each entry in *tables* is a dict with:
      'name'    – table name (goes into row 1 of the block)
      'headers' – list of header labels for row 2 of the block
      'rows'    – list of row-dicts {header_label: cell_value}
    Blocks are laid out side-by-side with one blank separator column between
    them (matching the format shown in the screenshot).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'tables_config_v2'

    col_offset = 0
    for tbl in tables:
        headers = tbl['headers']
        # row 1: table name in the first column of this block
        ws.cell(row=1, column=col_offset + 1, value=tbl['name'])
        # row 2: headers
        for h_idx, header in enumerate(headers):
            ws.cell(row=2, column=col_offset + 1 + h_idx, value=header)
        # rows 3+: data
        for r_idx, row_data in enumerate(tbl['rows']):
            for h_idx, header in enumerate(headers):
                val = row_data.get(header)
                if val is not None:
                    ws.cell(row=3 + r_idx, column=col_offset + 1 + h_idx, value=val)
        # blank separator column between blocks
        col_offset += len(headers) + 1

    payload = BytesIO()
    wb.save(payload)
    return payload.getvalue()


V2_HEADERS = [
    'Описание',
    'Код колонки в БД',
    'Тип',
    'Размерность',
    'Обязательность',
    'Уникальность',
    'Первичный ключ',
    'Внешний ключ',
    'Значение по умолчанию',
]


def test_parse_excel_v2_single_table():
    content = _make_v2_workbook([
        {
            'name': 'users',
            'headers': V2_HEADERS,
            'rows': [
                {'Описание': 'Идентификатор', 'Код колонки в БД': 'id', 'Тип': 'bigserial',
                 'Обязательность': 'да', 'Первичный ключ': 'да'},
                {'Описание': 'Имя', 'Код колонки в БД': 'name', 'Тип': 'varchar',
                 'Размерность': '100', 'Обязательность': 'да'},
            ],
        }
    ])

    tables = _parser.parse_tables_config(content, 'config.xlsm')

    assert len(tables) == 1
    assert tables[0].name == 'users'
    cols = {c.name: c for c in tables[0].columns}
    assert set(cols) == {'id', 'name'}
    assert cols['id'].primary_key is True
    assert cols['id'].nullable is False
    assert cols['name'].size == '100'
    assert cols['name'].label == 'Имя'


def test_parse_excel_v2_two_tables_side_by_side():
    content = _make_v2_workbook([
        {
            'name': 'orders',
            'headers': V2_HEADERS,
            'rows': [
                {'Код колонки в БД': 'id', 'Тип': 'bigserial', 'Первичный ключ': 'да'},
                {'Код колонки в БД': 'user_id', 'Тип': 'bigint',
                 'Внешний ключ': 'users(id)'},
            ],
        },
        {
            'name': 'products',
            'headers': V2_HEADERS,
            'rows': [
                {'Код колонки в БД': 'id', 'Тип': 'bigserial'},
                {'Код колонки в БД': 'title', 'Тип': 'varchar', 'Размерность': '255'},
            ],
        },
    ])

    tables = _parser.parse_tables_config(content, 'config.xlsm')

    assert len(tables) == 2
    names = {t.name for t in tables}
    assert names == {'orders', 'products'}

    orders = next(t for t in tables if t.name == 'orders')
    assert orders.columns[1].foreign_key == 'users(id)'

    products = next(t for t in tables if t.name == 'products')
    assert products.columns[1].size == '255'


def test_parse_excel_v2_and_v1_in_same_workbook():
    """A workbook with both sheets: tables from each sheet are combined."""
    wb = Workbook()

    # v1 sheet
    ws1 = wb.active
    ws1.title = 'tables_config'
    ws1['A1'] = 'Наименование таблицы'
    ws1['B1'] = 'v1_table'
    ws1['A3'] = 'Код колонки в БД'
    ws1['B3'] = 'id'
    ws1['A4'] = 'Тип'
    ws1['B4'] = 'bigserial'

    # v2 sheet
    ws2 = wb.create_sheet('tables_config_v2')
    ws2.cell(row=1, column=1, value='v2_table')
    ws2.cell(row=2, column=1, value='Описание')
    ws2.cell(row=2, column=2, value='Код колонки в БД')
    ws2.cell(row=2, column=3, value='Тип')
    ws2.cell(row=3, column=2, value='code')
    ws2.cell(row=3, column=3, value='varchar')

    payload = BytesIO()
    wb.save(payload)

    tables = _parser.parse_tables_config(payload.getvalue(), 'config.xlsm')

    table_names = {t.name for t in tables}
    assert 'v1_table' in table_names
    assert 'v2_table' in table_names


def test_parse_excel_v2_invalid_primary_key_raises_error():
    content = _make_v2_workbook([
        {
            'name': 'test_table',
            'headers': V2_HEADERS,
            'rows': [
                {'Код колонки в БД': 'id', 'Тип': 'bigserial', 'Первичный ключ': 'pk'},
            ],
        }
    ])

    with pytest.raises(AppError, match='Первичный ключ'):
        _parser.parse_tables_config(content, 'config.xlsm')


def test_parse_excel_v2_invalid_foreign_key_raises_error():
    content = _make_v2_workbook([
        {
            'name': 'test_table',
            'headers': V2_HEADERS,
            'rows': [
                {'Код колонки в БД': 'fk_col', 'Тип': 'bigint',
                 'Внешний ключ': 'other.id'},  # dot notation is invalid
            ],
        }
    ])

    with pytest.raises(AppError, match='некорректный формат ссылки'):
        _parser.parse_tables_config(content, 'config.xlsm')


def test_parse_excel_v2_table_name_label_in_row1():
    """Row 1 has 'Наименование таблицы' at start_col; actual name is at start_col+1."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'tables_config_v2'

    # Block 1: label at A1, actual table name at B1
    ws.cell(row=1, column=1, value='Наименование таблицы')
    ws.cell(row=1, column=2, value='users')
    ws.cell(row=2, column=1, value='Описание')
    ws.cell(row=2, column=2, value='Код колонки в БД')
    ws.cell(row=2, column=3, value='Тип')
    ws.cell(row=3, column=2, value='id')
    ws.cell(row=3, column=3, value='bigserial')

    payload = BytesIO()
    wb.save(payload)

    tables = _parser.parse_tables_config(payload.getvalue(), 'config.xlsm')

    assert len(tables) == 1
    assert tables[0].name == 'users'
    assert tables[0].columns[0].name == 'id'


def test_parse_excel_invalid_primary_key_value_raises_error():
    wb = Workbook()
    ws = wb.active
    ws.title = 'tables_config'

    ws['A1'] = 'Наименование таблицы'
    ws['B1'] = 'test_table'
    ws['A3'] = 'Код колонки в БД'
    ws['B3'] = 'id'
    ws['C3'] = 'code'
    ws['A4'] = 'Тип'
    ws['B4'] = 'bigserial'
    ws['C4'] = 'varchar'
    ws['A8'] = 'Первичный ключ'
    ws['B8'] = 'pk'  # invalid — must be "да", "нет", or empty

    payload = BytesIO()
    wb.save(payload)

    with pytest.raises(AppError, match='Первичный ключ'):
        _parser.parse_tables_config(payload.getvalue(), 'config.xlsx')


def test_parse_excel_da_net_for_required_and_unique():
    wb = Workbook()
    ws = wb.active
    ws.title = 'tables_config'

    ws['A1'] = 'Наименование таблицы'
    ws['B1'] = 'orders'
    ws['A3'] = 'Код колонки в БД'
    ws['B3'] = 'id'
    ws['C3'] = 'code'
    ws['D3'] = 'note'
    ws['A4'] = 'Тип'
    ws['B4'] = 'bigserial'
    ws['C4'] = 'varchar'
    ws['D4'] = 'text'
    ws['A6'] = 'Обязательность'
    ws['B6'] = 'да'   # required → nullable=False
    ws['C6'] = 'нет'  # not required → nullable=True
    # D6 empty → nullable=True (default)
    ws['A7'] = 'Уникальность'
    ws['B7'] = 'нет'  # not unique
    ws['C7'] = 'да'   # unique=True
    # D7 empty → unique=False (default)

    payload = BytesIO()
    wb.save(payload)

    tables = _parser.parse_tables_config(payload.getvalue(), 'config.xlsx')

    assert len(tables) == 1
    cols = {c.name: c for c in tables[0].columns}

    assert cols['id'].nullable is False   # Обязательность = "да"
    assert cols['id'].unique is False     # Уникальность = "нет"

    assert cols['code'].nullable is True  # Обязательность = "нет"
    assert cols['code'].unique is True    # Уникальность = "да"

    assert cols['note'].nullable is True  # Обязательность empty
    assert cols['note'].unique is False   # Уникальность empty


def test_parse_excel_invalid_required_value_raises_error():
    wb = Workbook()
    ws = wb.active
    ws.title = 'tables_config'

    ws['A1'] = 'Наименование таблицы'
    ws['B1'] = 'test_table'
    ws['A3'] = 'Код колонки в БД'
    ws['B3'] = 'id'
    ws['A4'] = 'Тип'
    ws['B4'] = 'bigserial'
    ws['A6'] = 'Обязательность'
    ws['B6'] = 'not null'  # invalid — must be "да", "нет", or empty

    payload = BytesIO()
    wb.save(payload)

    with pytest.raises(AppError, match='Обязательность'):
        _parser.parse_tables_config(payload.getvalue(), 'config.xlsx')


def test_parse_excel_invalid_unique_value_raises_error():
    wb = Workbook()
    ws = wb.active
    ws.title = 'tables_config'

    ws['A1'] = 'Наименование таблицы'
    ws['B1'] = 'test_table'
    ws['A3'] = 'Код колонки в БД'
    ws['B3'] = 'id'
    ws['A4'] = 'Тип'
    ws['B4'] = 'bigserial'
    ws['A7'] = 'Уникальность'
    ws['B7'] = 'unique'  # invalid — must be "да", "нет", or empty

    payload = BytesIO()
    wb.save(payload)

    with pytest.raises(AppError, match='Уникальность'):
        _parser.parse_tables_config(payload.getvalue(), 'config.xlsx')


def test_parse_excel_invalid_reference_format_raises_error():
    wb = Workbook()
    ws = wb.active
    ws.title = 'tables_config'

    ws['A1'] = 'Наименование таблицы'
    ws['B1'] = 'test_table'
    ws['A3'] = 'Код колонки в БД'
    ws['B3'] = 'id'
    ws['C3'] = 'fk_col'
    ws['A4'] = 'Тип'
    ws['B4'] = 'bigserial'
    ws['C4'] = 'bigint'
    ws['A8'] = 'Первичный ключ'
    ws['A9'] = 'Внешний ключ'
    ws['C9'] = 'other_table.id'  # invalid — must be other_table(id)

    payload = BytesIO()
    wb.save(payload)

    with pytest.raises(AppError, match='некорректный формат ссылки'):
        _parser.parse_tables_config(payload.getvalue(), 'config.xlsx')


def test_parse_excel_v1_default_value():
    wb = Workbook()
    ws = wb.active
    ws.title = 'tables_config'

    ws['A1'] = 'Наименование таблицы'
    ws['B1'] = 'settings'
    ws['A3'] = 'Код колонки в БД'
    ws['B3'] = 'status'
    ws['C3'] = 'score'
    ws['A4'] = 'Тип'
    ws['B4'] = 'varchar'
    ws['C4'] = 'integer'
    ws['A5'] = 'Значение по умолчанию'
    ws['B5'] = 'active'
    ws['C5'] = '0'

    payload = BytesIO()
    wb.save(payload)

    tables = _parser.parse_tables_config(payload.getvalue(), 'config.xlsx')

    assert len(tables) == 1
    cols = {c.name: c for c in tables[0].columns}
    assert cols['status'].default == 'active'
    assert cols['score'].default == '0'


def test_parse_excel_v2_default_value():
    content = _make_v2_workbook([
        {
            'name': 'settings',
            'headers': V2_HEADERS,
            'rows': [
                {'Код колонки в БД': 'status', 'Тип': 'varchar',
                 'Значение по умолчанию': 'active'},
                {'Код колонки в БД': 'score', 'Тип': 'integer',
                 'Значение по умолчанию': '0'},
                {'Код колонки в БД': 'note', 'Тип': 'text'},
            ],
        }
    ])

    tables = _parser.parse_tables_config(content, 'config.xlsm')

    assert len(tables) == 1
    cols = {c.name: c for c in tables[0].columns}
    assert cols['status'].default == 'active'
    assert cols['score'].default == '0'
    assert cols['note'].default is None
