"""Tests for the table configuration generator."""

import tempfile
from pathlib import Path

import openpyxl
import pytest

from main import (
    ColumnConfig,
    TableConfig,
    build_column_details_sheet,
    build_tables_config_v2_sheet,
    generate_table_config,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def sample_tables():
    return [
        TableConfig(
            table_name="users",
            schema="public",
            comment="Application users",
            columns=[
                ColumnConfig("id", "BIGINT", nullable=False, primary_key=True),
                ColumnConfig("email", "VARCHAR(255)", nullable=False),
            ],
        ),
        TableConfig(
            table_name="orders",
            schema="public",
            columns=[
                ColumnConfig("id", "BIGINT", nullable=False, primary_key=True),
                ColumnConfig("total", "NUMERIC(12,2)", nullable=False),
            ],
        ),
    ]


@pytest.fixture
def empty_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


# ---------------------------------------------------------------------------
# build_tables_config_v2_sheet
# ---------------------------------------------------------------------------

class TestBuildTablesConfigV2Sheet:
    def test_sheet_created(self, empty_workbook, sample_tables):
        build_tables_config_v2_sheet(empty_workbook, sample_tables)
        assert "tables_config_v2" in empty_workbook.sheetnames

    def test_header_row(self, empty_workbook, sample_tables):
        build_tables_config_v2_sheet(empty_workbook, sample_tables)
        ws = empty_workbook["tables_config_v2"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        assert headers == ["#", "Schema", "Table Name", "Comment", "Columns Count", "Has PK"]

    def test_data_rows(self, empty_workbook, sample_tables):
        build_tables_config_v2_sheet(empty_workbook, sample_tables)
        ws = empty_workbook["tables_config_v2"]
        assert ws.cell(row=2, column=3).value == "users"
        assert ws.cell(row=2, column=5).value == 2   # columns count
        assert ws.cell(row=2, column=6).value == "Yes"  # has PK

    def test_row_count_matches_tables(self, empty_workbook, sample_tables):
        build_tables_config_v2_sheet(empty_workbook, sample_tables)
        ws = empty_workbook["tables_config_v2"]
        # 1 header row + one row per table
        assert ws.max_row == 1 + len(sample_tables)

    def test_existing_sheet_replaced(self, empty_workbook, sample_tables):
        empty_workbook.create_sheet("tables_config_v2")
        build_tables_config_v2_sheet(empty_workbook, sample_tables)
        assert empty_workbook.sheetnames.count("tables_config_v2") == 1


# ---------------------------------------------------------------------------
# build_column_details_sheet
# ---------------------------------------------------------------------------

class TestBuildColumnDetailsSheet:
    def test_sheet_created(self, empty_workbook, sample_tables):
        build_column_details_sheet(empty_workbook, sample_tables[0])
        assert "users" in empty_workbook.sheetnames

    def test_title_cell(self, empty_workbook, sample_tables):
        build_column_details_sheet(empty_workbook, sample_tables[0])
        ws = empty_workbook["users"]
        assert ws["A1"].value == "public.users"

    def test_column_data(self, empty_workbook, sample_tables):
        build_column_details_sheet(empty_workbook, sample_tables[0])
        ws = empty_workbook["users"]
        # Find the header row (first row containing "#")
        header_row = next(r for r in ws.iter_rows(values_only=True) if r[0] == "#")
        assert "Column Name" in header_row

    def test_existing_sheet_replaced(self, empty_workbook, sample_tables):
        empty_workbook.create_sheet("users")
        build_column_details_sheet(empty_workbook, sample_tables[0])
        assert empty_workbook.sheetnames.count("users") == 1


# ---------------------------------------------------------------------------
# generate_table_config
# ---------------------------------------------------------------------------

class TestGenerateTableConfig:
    def test_file_created(self, sample_tables, tmp_path):
        out = tmp_path / "config.xlsx"
        generate_table_config(sample_tables, output_path=out)
        assert out.exists()

    def test_active_sheet_is_tables_config_v2(self, sample_tables, tmp_path):
        out = tmp_path / "config.xlsx"
        generate_table_config(sample_tables, output_path=out)
        wb = openpyxl.load_workbook(out)
        assert wb.active.title == "tables_config_v2"

    def test_all_table_sheets_present(self, sample_tables, tmp_path):
        out = tmp_path / "config.xlsx"
        generate_table_config(sample_tables, output_path=out)
        wb = openpyxl.load_workbook(out)
        for table in sample_tables:
            assert table.table_name in wb.sheetnames

    def test_tables_config_v2_sheet_present(self, sample_tables, tmp_path):
        out = tmp_path / "config.xlsx"
        generate_table_config(sample_tables, output_path=out)
        wb = openpyxl.load_workbook(out)
        assert "tables_config_v2" in wb.sheetnames

    def test_returns_path(self, sample_tables, tmp_path):
        out = tmp_path / "config.xlsx"
        result = generate_table_config(sample_tables, output_path=out)
        assert isinstance(result, Path)
        assert result == out

    def test_empty_tables_list(self, tmp_path):
        out = tmp_path / "empty.xlsx"
        generate_table_config([], output_path=out)
        wb = openpyxl.load_workbook(out)
        assert "tables_config_v2" in wb.sheetnames
        assert wb.active.title == "tables_config_v2"
