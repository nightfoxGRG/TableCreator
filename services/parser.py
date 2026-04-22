import json
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from services.models import ColumnConfig, ConfigParseError, TableConfig
from services.validators import _validate_reference_cell, _validate_yes_no_cell


TABLE_NAME_LABELS = {'наименование таблицы', 'table_name'}
COLUMN_CODE_LABELS = {'код колонки в бд', 'column_code'}
COLUMN_NAME_LABELS = {'описание', 'column_name'}
TYPE_LABELS = {'тип', 'type'}
SIZE_LABELS = {'размерность', 'size'}
REQUIRED_LABELS = {'обязательность', 'nullable'}
UNIQUE_LABELS = {'уникальность', 'unique'}
PRIMARY_KEY_LABELS = {'первичный ключ', 'primary_key'}
FOREIGN_KEY_LABELS = {'внешний ключ', 'foreign_key'}
DEFAULT_LABELS = {'значение по умолчанию', 'default'}


def parse_tables_config(content: bytes, filename: str) -> list[TableConfig]:
    extension = Path(filename).suffix.lower()
    if extension in {'.xlsx', '.xlsm'}:
        tables = _parse_excel(content)
    elif extension == '.json':
        data = json.loads(content.decode('utf-8'))
        tables = _parse_structured_tables(data)
    else:
        raise ConfigParseError('Неподдерживаемый формат файла.')

    if not tables:
        raise ConfigParseError('Секция tables_config не найдена или не содержит таблиц.')
    return tables


def _parse_structured_tables(data: dict | None) -> list[TableConfig]:
    if not isinstance(data, dict):
        raise ConfigParseError('Некорректная структура конфигурации.')

    tables_config = data.get('tables_config')
    if tables_config is None:
        raise ConfigParseError('Секция tables_config отсутствует.')

    tables: list[TableConfig] = []
    if isinstance(tables_config, dict):
        iterable = [
            {'table_name': table_name, 'columns': columns}
            for table_name, columns in tables_config.items()
        ]
    elif isinstance(tables_config, list):
        iterable = tables_config
    else:
        raise ConfigParseError('Секция tables_config должна быть словарем или списком.')

    for table_item in iterable:
        if not isinstance(table_item, dict):
            raise ConfigParseError('Каждая таблица в tables_config должна быть объектом.')
        table_name = _normalize_text(table_item.get('table_name'))
        if not table_name:
            raise ConfigParseError('У таблицы отсутствует имя.')
        columns_raw = table_item.get('columns', [])
        if isinstance(columns_raw, dict):
            columns_raw = [{'column_code': name, **details} for name, details in columns_raw.items()]
        if not isinstance(columns_raw, list):
            raise ConfigParseError(f'Колонки таблицы {table_name} должны быть списком или словарем.')

        columns: list[ColumnConfig] = []
        for column_item in columns_raw:
            if not isinstance(column_item, dict):
                raise ConfigParseError(f'Некорректное описание колонки в таблице {table_name}.')
            column_name = _normalize_text(column_item.get('column_code'))
            db_type = _normalize_text(column_item.get('type'))
            if not column_name or not db_type:
                raise ConfigParseError(f'Колонка в таблице {table_name} должна содержать column_code и type.')

            columns.append(
                ColumnConfig(
                    name=column_name,
                    db_type=db_type,
                    size=_normalize_text(column_item.get('size')),
                    nullable=_to_bool(column_item.get('nullable', True), default=True),
                    unique=_to_bool(column_item.get('unique', False), default=False),
                    primary_key=_to_bool(column_item.get('primary_key', False), default=False),
                    foreign_key=_normalize_text(column_item.get('foreign_key')),
                    default=_normalize_text(column_item.get('default')),
                    label=_normalize_text(column_item.get('column_name')),
                )
            )

        tables.append(TableConfig(name=table_name, columns=columns))

    return tables


def _parse_excel(content: bytes) -> list[TableConfig]:
    workbook = load_workbook(BytesIO(content), data_only=True)

    v1_sheet = None
    v2_sheet = None
    for sheet in workbook.worksheets:
        name = sheet.title.strip().lower()
        if name == 'tables_config':
            v1_sheet = sheet
        elif name == 'tables_config_v2':
            v2_sheet = sheet

    if v1_sheet is None and v2_sheet is None:
        if len(workbook.worksheets) == 1:
            sheet = workbook.worksheets[0]
            if _is_v2_format(sheet):
                v2_sheet = sheet
            else:
                v1_sheet = sheet
        else:
            raise ConfigParseError('Лист tables_config не найден.')

    tables: list[TableConfig] = []

    if v1_sheet is not None:
        rows = [list(row) for row in v1_sheet.iter_rows(values_only=True)]
        tables.extend(_parse_excel_v1_rows(rows))

    if v2_sheet is not None:
        tables.extend(_parse_excel_v2_sheet(v2_sheet))

    return tables


def _is_v2_format(sheet) -> bool:
    rows = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    if not rows:
        return False
    for cell in rows[0]:
        label = _label(cell)
        if label in COLUMN_NAME_LABELS or label in COLUMN_CODE_LABELS:
            return True
    return False


def _parse_excel_v1_rows(rows: list[list]) -> list[TableConfig]:
    tables: list[TableConfig] = []
    row_index = 0
    while row_index < len(rows):
        current_label = _label(rows[row_index][0] if rows[row_index] else None)
        if current_label in TABLE_NAME_LABELS:
            table, row_index = _parse_excel_table_block(rows, row_index)
            if table is not None:
                tables.append(table)
            continue
        row_index += 1
    return tables


def _parse_excel_v2_sheet(sheet) -> list[TableConfig]:
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    if len(rows) < 2:
        return []

    table_row = rows[0]   # row 1: table name for each block
    header_row = rows[1]  # row 2: column-attribute headers

    # Each table block starts at the column where a COLUMN_NAME_LABELS header appears
    block_starts: list[int] = []
    for col_idx, cell in enumerate(header_row):
        if _label(cell) in COLUMN_NAME_LABELS:
            block_starts.append(col_idx)

    tables: list[TableConfig] = []
    for i, start_col in enumerate(block_starts):
        raw_name_cell = table_row[start_col] if start_col < len(table_row) else None
        # If row-1 cell at start_col is the "Наименование таблицы" label rather than
        # the actual table name, the real name is in the adjacent cell to the right.
        if _label(raw_name_cell) in TABLE_NAME_LABELS:
            raw_name_cell = table_row[start_col + 1] if start_col + 1 < len(table_row) else None
        table_name = _normalize_text(raw_name_cell)
        if not table_name:
            continue

        end_col = block_starts[i + 1] if i + 1 < len(block_starts) else len(header_row)

        # Map label -> column index for this block
        header_map: dict[str, int] = {}
        for col_idx in range(start_col, min(end_col, len(header_row))):
            label = _label(header_row[col_idx])
            if label and label not in header_map:
                header_map[label] = col_idx

        code_col = _find_col(header_map, COLUMN_CODE_LABELS)
        type_col = _find_col(header_map, TYPE_LABELS)
        name_col = _find_col(header_map, COLUMN_NAME_LABELS)
        size_col = _find_col(header_map, SIZE_LABELS)
        required_col = _find_col(header_map, REQUIRED_LABELS)
        unique_col = _find_col(header_map, UNIQUE_LABELS)
        pk_col = _find_col(header_map, PRIMARY_KEY_LABELS)
        fk_col = _find_col(header_map, FOREIGN_KEY_LABELS)
        default_col = _find_col(header_map, DEFAULT_LABELS)

        if code_col is None or type_col is None:
            raise ConfigParseError(
                f'Для таблицы {table_name} отсутствуют заголовки кодов колонок или типов.'
            )

        columns: list[ColumnConfig] = []
        for row in rows[2:]:
            code = _cell(row, code_col)
            if not code:
                continue

            db_type = _cell(row, type_col)
            if not db_type:
                raise ConfigParseError(
                    f'У колонки {code} таблицы {table_name} не указан тип.'
                )

            fk_value = _cell(row, fk_col) if fk_col is not None else None
            req_val = _cell(row, required_col) if required_col is not None else None
            uniq_val = _cell(row, unique_col) if unique_col is not None else None
            pk_val = _cell(row, pk_col) if pk_col is not None else None

            _validate_yes_no_cell(req_val, 'Обязательность', code, table_name)
            _validate_yes_no_cell(uniq_val, 'Уникальность', code, table_name)
            _validate_yes_no_cell(pk_val, 'Первичный ключ', code, table_name)
            _validate_reference_cell(fk_value, code, table_name)

            columns.append(
                ColumnConfig(
                    name=code,
                    db_type=db_type,
                    size=_cell(row, size_col) if size_col is not None else None,
                    nullable=not _contains_any(req_val, {'да'}),
                    unique=_contains_any(uniq_val, {'да'}),
                    primary_key=_contains_any(pk_val, {'да'}),
                    foreign_key=fk_value,
                    default=_cell(row, default_col) if default_col is not None else None,
                    label=_cell(row, name_col) if name_col is not None else None,
                )
            )

        tables.append(TableConfig(name=table_name, columns=columns))

    return tables


def _find_tables_sheet(workbook):
    for sheet in workbook.worksheets:
        if sheet.title.strip().lower() == 'tables_config':
            return sheet
    if len(workbook.worksheets) == 1:
        return workbook.worksheets[0]
    raise ConfigParseError('Лист tables_config не найден.')


def _parse_excel_table_block(rows: list[list], start_index: int) -> tuple[TableConfig | None, int]:
    header_row = rows[start_index]
    table_name = _first_non_empty(header_row[1:])
    if not table_name:
        return None, start_index + 1

    end_index = start_index + 1
    row_map: dict[str, list] = {}
    while end_index < len(rows):
        row = rows[end_index]
        first_value = row[0] if row else None
        label = _label(first_value)
        if label in TABLE_NAME_LABELS:
            break
        if label:
            row_map[label] = row
        end_index += 1

    code_row = _find_row(row_map, COLUMN_CODE_LABELS)
    type_row = _find_row(row_map, TYPE_LABELS)
    name_row = _find_row(row_map, COLUMN_NAME_LABELS)
    size_row = _find_row(row_map, SIZE_LABELS)
    required_row = _find_row(row_map, REQUIRED_LABELS)
    unique_row = _find_row(row_map, UNIQUE_LABELS)
    primary_key_row = _find_row(row_map, PRIMARY_KEY_LABELS)
    foreign_key_row = _find_row(row_map, FOREIGN_KEY_LABELS)
    default_row = _find_row(row_map, DEFAULT_LABELS)

    if not code_row or not type_row:
        raise ConfigParseError(f'Для таблицы {table_name} отсутствуют строки с кодами колонок или типами.')

    max_len = max(
        len(code_row),
        len(type_row),
        len(name_row or []),
        len(size_row or []),
        len(required_row or []),
        len(unique_row or []),
        len(primary_key_row or []),
        len(foreign_key_row or []),
        len(default_row or []),
    )

    columns: list[ColumnConfig] = []
    for idx in range(1, max_len):
        name = _cell(code_row, idx)
        db_type = _cell(type_row, idx)
        if not name:
            continue
        if not db_type:
            raise ConfigParseError(f'У колонки {name} таблицы {table_name} не указан тип.')

        foreign_key_value = _cell(foreign_key_row, idx)

        _validate_yes_no_cell(_cell(required_row, idx), 'Обязательность', name, table_name)
        _validate_yes_no_cell(_cell(unique_row, idx), 'Уникальность', name, table_name)
        _validate_yes_no_cell(_cell(primary_key_row, idx), 'Первичный ключ', name, table_name)
        _validate_reference_cell(foreign_key_value, name, table_name)

        columns.append(
            ColumnConfig(
                name=name,
                db_type=db_type,
                size=_cell(size_row, idx),
                nullable=not _contains_any(_cell(required_row, idx), {'да'}),
                unique=_contains_any(_cell(unique_row, idx), {'да'}),
                primary_key=_contains_any(_cell(primary_key_row, idx), {'да'}),
                foreign_key=foreign_key_value,
                default=_cell(default_row, idx),
                label=_cell(name_row, idx),
            )
        )

    return TableConfig(name=table_name, columns=columns), end_index


def _find_row(row_map: dict[str, list], names: set[str]) -> list | None:
    for name in names:
        if name in row_map:
            return row_map[name]
    return None


def _find_col(header_map: dict[str, int], labels: set[str]) -> int | None:
    for label in labels:
        if label in header_map:
            return header_map[label]
    return None


def _cell(row: list | None, index: int) -> str | None:
    if not row or index >= len(row):
        return None
    return _normalize_text(row[index])


def _label(value) -> str:
    return _normalize_text(value, lower=True) or ''


def _normalize_text(value, lower: bool = False) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return text.lower() if lower else text


def _first_non_empty(values: list) -> str | None:
    for value in values:
        text = _normalize_text(value)
        if text:
            return text
    return None


def _contains_any(value: str | None, expected_values: set[str]) -> bool:
    if not value:
        return False
    lower_value = value.lower()
    return any(item in lower_value for item in expected_values)


def _to_bool(value, default: bool) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    normalized = str(value).strip().lower()
    if normalized in {'1', 'true', 'yes', 'y', 'да'}:
        return True
    if normalized in {'0', 'false', 'no', 'n', 'нет'}:
        return False
    return default
