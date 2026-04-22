"""Generate an Excel configuration file in tables_config_v2 format
from a list of inferred column descriptors.
"""

import re
import zipfile
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


_TEMPLATE_PATH = Path(__file__).parent.parent / 'static' / 'TablesConfig.xlsm'

# Number of data columns in one table block (matches the template header row)
_V2_DATA_COLS = 9


def _sheet_name_to_zip_path(workbook_xml: str, rels_xml: str) -> dict[str, str]:
    """Return {sheet_name: zip_entry_path} parsed from workbook.xml and its .rels.

    Handles both absolute targets (/xl/worksheets/sheet1.xml) and relative targets
    (worksheets/sheet1.xml) that are resolved relative to the xl/ directory.
    """
    name_rid: dict[str, str] = {}
    for m in re.finditer(r'<sheet\s[^>]*\bname="([^"]*)"[^>]*\br:id="([^"]*)"', workbook_xml):
        name_rid[m.group(1)] = m.group(2)
    rid_zip: dict[str, str] = {}
    for m in re.finditer(r'<Relationship\s([^>]*)>', rels_xml):
        attrs = m.group(1)
        id_m = re.search(r'\bId="([^"]*)"', attrs)
        target_m = re.search(r'\bTarget="([^"]*)"', attrs)
        if not id_m or not target_m:
            continue
        target = target_m.group(1)
        if target.startswith('/'):
            zip_path = target.lstrip('/')      # /xl/worksheets/sheet1.xml → xl/worksheets/sheet1.xml
        else:
            zip_path = f'xl/{target}'          # worksheets/sheet1.xml → xl/worksheets/sheet1.xml
        rid_zip[id_m.group(1)] = zip_path
    return {name: rid_zip[rid] for name, rid in name_rid.items() if rid in rid_zip}


def _restore_x14_validations(template_bytes: bytes, output_bytes: bytes) -> bytes:
    """Re-inject <extLst> blocks (containing x14:dataValidations) that openpyxl drops.

    openpyxl does not support Excel 2010+ extended data validations (x14:dataValidation
    inside <extLst>) and silently removes them on save.  This function copies those
    blocks verbatim from *template_bytes* into the matching worksheets of *output_bytes*.
    """
    wb_rel_path = 'xl/_rels/workbook.xml.rels'
    wb_path = 'xl/workbook.xml'

    with zipfile.ZipFile(BytesIO(template_bytes)) as tz:
        tmpl_name_to_zip = _sheet_name_to_zip_path(
            tz.read(wb_path).decode('utf-8'),
            tz.read(wb_rel_path).decode('utf-8'),
        )
        # Map sheet name -> (extLst XML string, extra namespace declarations needed)
        tmpl_extlst: dict[str, tuple[str, str] | None] = {}
        for name, zip_path in tmpl_name_to_zip.items():
            xml = tz.read(zip_path).decode('utf-8')
            m = re.search(r'<extLst>.*?</extLst>', xml, re.DOTALL)
            if not m:
                tmpl_extlst[name] = None
                continue
            extlst = m.group(0)
            # Collect namespace declarations from the template worksheet root so we can
            # carry any that the extLst needs but openpyxl omits (e.g. xmlns:xr).
            root_m = re.search(r'<worksheet([^>]*)>', xml)
            root_attrs = root_m.group(1) if root_m else ''
            tmpl_extlst[name] = (extlst, root_attrs)

    out_buf = BytesIO(output_bytes)
    with zipfile.ZipFile(out_buf, 'r') as oz:
        out_name_to_zip = _sheet_name_to_zip_path(
            oz.read(wb_path).decode('utf-8'),
            oz.read(wb_rel_path).decode('utf-8'),
        )
        out_entries = {item.filename: oz.read(item.filename) for item in oz.infolist()}

    for sheet_name, payload in tmpl_extlst.items():
        if not payload:
            continue
        extlst, tmpl_root_attrs = payload
        zip_path = out_name_to_zip.get(sheet_name)
        if not zip_path or zip_path not in out_entries:
            continue
        xml = out_entries[zip_path].decode('utf-8')
        # Ensure any namespace prefixes used by the extLst are declared on the root element
        root_tag_m = re.match(r'(<worksheet)([^>]*)(>)', xml)
        if root_tag_m:
            cur_attrs = root_tag_m.group(2)
            extra_ns = ''
            for ns_m in re.finditer(r'(xmlns:[a-zA-Z0-9_]+="[^"]*")', tmpl_root_attrs):
                ns_decl = ns_m.group(1)
                prefix = re.match(r'xmlns:([a-zA-Z0-9_]+)=', ns_decl).group(1)
                if (f'{prefix}:' in extlst) and (f'xmlns:{prefix}=' not in cur_attrs):
                    extra_ns += ' ' + ns_decl
            if extra_ns:
                xml = root_tag_m.group(1) + cur_attrs + extra_ns + root_tag_m.group(3) + xml[root_tag_m.end():]
        # Remove any existing extLst the output may already carry, then re-add template's
        xml = re.sub(r'<extLst>.*?</extLst>', '', xml, flags=re.DOTALL)
        xml = xml.replace('</worksheet>', extlst + '</worksheet>', 1)
        out_entries[zip_path] = xml.encode('utf-8')

    result = BytesIO()
    with zipfile.ZipFile(result, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
        for filename, data in out_entries.items():
            zout.writestr(filename, data)
    return result.getvalue()


def generate_excel_config_v2(table_name: str, columns: list[dict]) -> bytes:
    """Build and return bytes of an xlsm workbook containing a tables_config_v2 sheet.

    Loads TablesConfig.xlsm as a template and fills in the first table block
    without deleting any rows or overwriting styles, formulas, or other settings.
    Only cell values are written.

    Each dict in *columns* must contain at least:
      'code'    – column code (SQL identifier)
      'db_type' – PostgreSQL type
    Optional keys: 'label', 'size'.
    """
    wb = load_workbook(_TEMPLATE_PATH, keep_vba=True)
    template_bytes = _TEMPLATE_PATH.read_bytes()

    ws = wb['tables_config_v2']
    wb.active = ws

    # ----- Write table name into B1 (A1 already holds "Наименование таблицы") -----
    ws.cell(row=1, column=2).value = table_name

    # ----- Clear old sample data values in block-1 data rows (rows 3+, cols A-I) -----
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=_V2_DATA_COLS):
        for cell in row:
            cell.value = None

    # ----- Write column data starting at row 3 -----
    for row_idx, col_info in enumerate(columns, start=3):
        ws.cell(row=row_idx, column=1).value = col_info.get('label') or col_info['code']
        ws.cell(row=row_idx, column=2).value = col_info['code']
        ws.cell(row=row_idx, column=3).value = col_info['db_type']
        size = col_info.get('size')
        if size:
            ws.cell(row=row_idx, column=4).value = size

    output = BytesIO()
    wb.save(output)
    return _restore_x14_validations(template_bytes, output.getvalue())
