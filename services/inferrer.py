"""Infer PostgreSQL column types and sizes from raw data file contents.

Supports reading Excel (.xlsx, .xlsm) and CSV files.
The first row of the input file is expected to contain column headers.
"""

import csv
import re
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook
from transliterate import translit
from transliterate.exceptions import LanguageDetectionError


# Patterns for date / datetime detection
_DATE_RE = re.compile(r'^\d{4}-\d{2}-\d{2}$')
_DATETIME_RE = re.compile(r'^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}')

# Strict boolean literals recognised in data
_BOOL_VALUES: frozenset[str] = frozenset({'true', 'false'})

# Extensions accepted by the upload endpoint for this feature
ALLOWED_DATA_EXTENSIONS = {'.xlsx', '.xlsm', '.csv'}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def read_data_file(content: bytes, filename: str) -> tuple[str, list[str], list[list]]:
    """Parse *content* from *filename* and return (table_name, headers, rows).

    *table_name* is derived from the filename stem.
    *headers* is the list of column names from the first row.
    *rows* is a list of value lists for every subsequent row.
    """
    extension = Path(filename).suffix.lower()
    stem = Path(filename).stem

    if extension in {'.xlsx', '.xlsm'}:
        headers, rows = _read_excel(content)
    elif extension == '.csv':
        headers, rows = _read_csv(content)
    else:
        from services.models import ConfigParseError
        raise ConfigParseError(
            f'Неподдерживаемый формат файла данных. '
            f'Допустимы: {", ".join(sorted(ALLOWED_DATA_EXTENSIONS))}.'
        )

    return stem, headers, rows


def infer_columns(headers: list[str], rows: list[list]) -> list[dict]:
    """Return a list of column-info dicts inferred from *headers* and *rows*.

    Each dict contains:
      'code'    – sanitized SQL identifier (derived from header)
      'label'   – original header string
      'db_type' – PostgreSQL type name
      'size'    – size/precision string or None
    """
    columns = []
    for col_idx, header in enumerate(headers):
        values = [row[col_idx] if col_idx < len(row) else None for row in rows]
        db_type, size = _infer_db_type(values)
        columns.append({
            'code': _sanitize_code(header),
            'label': header,
            'db_type': db_type,
            'size': size,
        })
    return columns


# ---------------------------------------------------------------------------
# Internal helpers — file reading
# ---------------------------------------------------------------------------

def _read_excel(content: bytes) -> tuple[list[str], list[list]]:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
    if not all_rows:
        from services.models import ConfigParseError
        raise ConfigParseError('Файл данных пустой.')
    headers = [str(cell) if cell is not None else '' for cell in all_rows[0]]
    rows = all_rows[1:]
    return headers, rows


def _read_csv(content: bytes) -> tuple[list[str], list[list]]:
    for encoding in ('utf-8-sig', 'utf-8', 'cp1251', 'latin-1'):
        try:
            text = content.decode(encoding)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    else:
        text = content.decode('latin-1')

    reader = csv.reader(StringIO(text))
    all_rows = list(reader)
    if not all_rows:
        from services.models import ConfigParseError
        raise ConfigParseError('Файл данных пустой.')
    headers = all_rows[0]
    rows = all_rows[1:]
    return headers, rows


# ---------------------------------------------------------------------------
# Internal helpers — type inference
# ---------------------------------------------------------------------------

def _non_null(values: list) -> list:
    return [v for v in values if v is not None and str(v).strip() != '']


def _try_int(v) -> bool:
    try:
        int(str(v).strip())
        return True
    except (ValueError, TypeError):
        return False


def _try_float(v) -> bool:
    try:
        float(str(v).strip())
        return True
    except (ValueError, TypeError):
        return False


def _is_bool(v) -> bool:
    return str(v).strip().lower() in _BOOL_VALUES


def _is_date(v) -> bool:
    return bool(_DATE_RE.match(str(v).strip()))


def _is_datetime(v) -> bool:
    return bool(_DATETIME_RE.match(str(v).strip()))


def _round_up_to_50(n: int) -> int:
    return max(50, ((n + 49) // 50) * 50)


def _infer_db_type(values: list) -> tuple[str, str | None]:
    """Return (db_type, size_or_None) for a column with the given *values*."""
    nn = _non_null(values)

    if not nn:
        return 'text', None

    # Boolean
    if all(_is_bool(v) for v in nn):
        return 'boolean', None

    # Date / datetime
    if all(_is_date(v) or _is_datetime(v) for v in nn):
        if any(_is_datetime(v) for v in nn):
            return 'timestamp', None
        return 'date', None

    # Integer
    if all(_try_int(v) for v in nn):
        max_val = max(abs(int(str(v).strip())) for v in nn)
        if max_val > 2_147_483_647:
            return 'bigint', None
        return 'integer', None

    # Numeric / float
    if all(_try_float(v) for v in nn):
        max_int_digits = 0
        max_dec_digits = 0
        for v in nn:
            s = str(float(str(v).strip()))
            if '.' in s:
                int_part, dec_part = s.split('.')
                max_int_digits = max(max_int_digits, len(int_part.lstrip('-')))
                dec_stripped = dec_part.rstrip('0')
                max_dec_digits = max(max_dec_digits, len(dec_stripped))
            else:
                max_int_digits = max(max_int_digits, len(s.lstrip('-')))
        if max_dec_digits > 0:
            precision = max_int_digits + max_dec_digits
            return 'numeric', f'{precision},{max_dec_digits}'
        return 'numeric', None

    # String
    max_len = max(len(str(v)) for v in nn)
    if max_len > 255:
        return 'text', None
    return 'varchar', str(_round_up_to_50(max_len))


# ---------------------------------------------------------------------------
# Internal helpers — identifier sanitization
# ---------------------------------------------------------------------------

def _transliterate_to_latin(name: str) -> str:
    """Return *name* with any Cyrillic (Russian) characters transliterated to Latin.

    Non-Cyrillic text is returned as-is.  If the transliteration library
    cannot detect the language, the original string is returned unchanged.
    """
    try:
        return translit(name, 'ru', reversed=True)
    except LanguageDetectionError:
        return name


def _sanitize_code(name: str) -> str:
    """Convert an arbitrary column header into a valid PostgreSQL identifier.

    Cyrillic text is first transliterated to Latin so that Russian column
    descriptions produce readable identifiers instead of falling back to 'col'.
    """
    code = _transliterate_to_latin(name).strip()
    code = re.sub(r"[\s\-']+", '_', code)
    code = re.sub(r'[^A-Za-z0-9_]', '', code)
    if code and code[0].isdigit():
        code = '_' + code
    return code.lower() or 'col'
