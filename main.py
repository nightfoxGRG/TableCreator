"""
Table Configuration Generator.

Generates an Excel workbook with table configuration sheets.
The active sheet in the produced file is set to 'tables_config_v2'.
"""

from __future__ import annotations

import dataclasses
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclasses.dataclass
class ColumnConfig:
    name: str
    data_type: str
    nullable: bool = True
    primary_key: bool = False
    default: str = ""
    comment: str = ""


@dataclasses.dataclass
class TableConfig:
    table_name: str
    schema: str = "public"
    columns: List[ColumnConfig] = dataclasses.field(default_factory=list)
    comment: str = ""


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(fill_type="solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
SUBHEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")


def _style_header_row(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def build_tables_config_v2_sheet(wb: openpyxl.Workbook, tables: List[TableConfig]) -> None:
    """Create (or replace) the 'tables_config_v2' sheet with a summary of all tables."""
    sheet_name = "tables_config_v2"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    headers = ["#", "Schema", "Table Name", "Comment", "Columns Count", "Has PK"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for idx, table in enumerate(tables, start=1):
        has_pk = any(c.primary_key for c in table.columns)
        ws.append([
            idx,
            table.schema,
            table.table_name,
            table.comment,
            len(table.columns),
            "Yes" if has_pk else "No",
        ])

    _auto_width(ws)


def build_column_details_sheet(wb: openpyxl.Workbook, table: TableConfig) -> None:
    """Create a dedicated sheet for each table's column definitions."""
    sheet_name = table.table_name[:31]  # Excel sheet name limit
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # Table header
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"{table.schema}.{table.table_name}"
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = CENTER
    title_cell.fill = SUBHEADER_FILL

    if table.comment:
        ws.merge_cells("A2:F2")
        ws["A2"].value = table.comment
        ws["A2"].alignment = CENTER

    headers = ["#", "Column Name", "Data Type", "Nullable", "Primary Key", "Default", "Comment"]
    ws.append(headers)
    _style_header_row(ws, ws.max_row, len(headers))

    for idx, col in enumerate(table.columns, start=1):
        ws.append([
            idx,
            col.name,
            col.data_type,
            "Yes" if col.nullable else "No",
            "Yes" if col.primary_key else "No",
            col.default,
            col.comment,
        ])

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Generator
# ---------------------------------------------------------------------------

def generate_table_config(tables: List[TableConfig], output_path: str | Path = "table_config.xlsx") -> Path:
    """
    Generate an Excel workbook from *tables* and save it to *output_path*.

    The active sheet in the resulting file is 'tables_config_v2'.
    """
    output_path = Path(output_path)
    wb = openpyxl.Workbook()

    # Remove the default empty sheet created by openpyxl
    default_sheet = wb.active
    wb.remove(default_sheet)

    # One sheet per table (column details)
    for table in tables:
        build_column_details_sheet(wb, table)

    # Summary sheet
    build_tables_config_v2_sheet(wb, tables)

    # Activate the 'tables_config_v2' sheet so it is selected when the file opens
    wb.active = wb["tables_config_v2"]

    wb.save(output_path)
    print(f"Configuration saved to '{output_path}' (active sheet: tables_config_v2)")
    return output_path


# ---------------------------------------------------------------------------
# Example usage
# ---------------------------------------------------------------------------

def _example_tables() -> List[TableConfig]:
    return [
        TableConfig(
            table_name="users",
            schema="public",
            comment="Application users",
            columns=[
                ColumnConfig("id", "BIGINT", nullable=False, primary_key=True, comment="Primary key"),
                ColumnConfig("username", "VARCHAR(100)", nullable=False, comment="Login name"),
                ColumnConfig("email", "VARCHAR(255)", nullable=False, comment="Email address"),
                ColumnConfig("created_at", "TIMESTAMP", nullable=False, default="NOW()", comment="Record creation time"),
                ColumnConfig("is_active", "BOOLEAN", nullable=False, default="TRUE"),
            ],
        ),
        TableConfig(
            table_name="orders",
            schema="public",
            comment="Customer orders",
            columns=[
                ColumnConfig("id", "BIGINT", nullable=False, primary_key=True),
                ColumnConfig("user_id", "BIGINT", nullable=False, comment="FK → users.id"),
                ColumnConfig("total_amount", "NUMERIC(12,2)", nullable=False),
                ColumnConfig("status", "VARCHAR(50)", nullable=False, default="'pending'"),
                ColumnConfig("created_at", "TIMESTAMP", nullable=False, default="NOW()"),
            ],
        ),
        TableConfig(
            table_name="products",
            schema="catalog",
            comment="Product catalogue",
            columns=[
                ColumnConfig("id", "BIGINT", nullable=False, primary_key=True),
                ColumnConfig("sku", "VARCHAR(64)", nullable=False, comment="Stock-keeping unit"),
                ColumnConfig("name", "VARCHAR(255)", nullable=False),
                ColumnConfig("price", "NUMERIC(10,2)", nullable=False),
                ColumnConfig("stock_qty", "INTEGER", nullable=False, default="0"),
            ],
        ),
    ]


if __name__ == "__main__":
    generate_table_config(_example_tables())
